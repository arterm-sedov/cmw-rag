"""CMW Platform Document Processor.

Processes document content (PDF, DOCX, XLSX, ZIP) and extracts text.
"""

import base64
import io
import os
import tempfile
import zipfile
from typing import Any

DEFAULT_PLATFORM = "primary"


def decode_base64_content(base64_string: str) -> dict[str, Any]:
    """Decode base64 string to bytes.

    Args:
        base64_string: Base64-encoded content

    Returns:
        Dict with keys:
            - success: bool
            - data: bytes (if success)
            - error: str (if failed)
    """
    try:
        data = base64.b64decode(base64_string)
        return {"success": True, "data": data}
    except Exception as e:
        return {"success": False, "error": f"Failed to decode base64: {e}"}


def detect_mime_type(data: bytes) -> str:
    """Detect MIME type from file magic bytes.

    Args:
        data: Raw file bytes

    Returns:
        MIME type string
    """
    magic_signatures = {
        b"%PDF": "application/pdf",
        b"PK\x03\x04": "application/vnd.openxmlformats-officedocument",  # DOCX/XLSX/ZIP
        b"\xd0\xcf\x11\xe0": "application/msword",  # DOC/XLS old format
    }

    for sig, mime in magic_signatures.items():
        if data.startswith(sig):
            return mime

    return "application/octet-stream"


def process_document(
    base64_content: str,
    mime_type: str | None = None,
    filename: str | None = None,
) -> dict[str, Any]:
    """Process document and extract text content.

    Supports: PDF, DOCX, XLSX, ZIP, images

    Args:
        base64_content: Base64-encoded document content
        mime_type: MIME type hint (detected from magic bytes if not provided)
        filename: Original filename for logging

    Returns:
        Dict with keys:
            - success: bool
            - text: extracted text (if success)
            - page_count: number of pages (if applicable)
            - file_type: file type string
            - error: str (if failed)
    """
    # Decode base64
    decoded = decode_base64_content(base64_content)
    if not decoded["success"]:
        return {"success": False, "error": decoded["error"]}

    data = decoded["data"]

    # Detect MIME type from content if not provided
    detected_mime = mime_type or detect_mime_type(data)

    # Route to appropriate processor
    if detected_mime == "application/pdf":
        return _process_pdf(data)

    if "wordprocessing" in detected_mime or "document" in detected_mime:
        return _process_docx(data)

    if "spreadsheet" in detected_mime or "excel" in detected_mime:
        return _process_xlsx(data)

    if detected_mime == "application/zip":
        return _process_zip(data)

    if detected_mime.startswith("image/"):
        return _process_image(data)

    return {"success": False, "error": f"Unsupported file type: {detected_mime}"}


def _process_pdf(data: bytes) -> dict[str, Any]:
    """Process PDF using PyMuPDF4LLM."""
    import pymupdf4llm

    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
        f.write(data)
        temp_path = f.name

    try:
        md = pymupdf4llm.to_markdown(
            temp_path,
            ignore_images=True,
            ignore_graphics=True,
            page_chunks=False,
        )
        return {
            "success": True,
            "text": md,
            "page_count": 1,
            "file_type": "pdf",
        }
    except ImportError:
        return {"success": False, "error": "PyMuPDF4LLM not installed"}
    except Exception as e:
        return {"success": False, "error": f"PDF processing failed: {e}"}
    finally:
        os.unlink(temp_path)


def _process_docx(data: bytes) -> dict[str, Any]:
    """Process Word document using direct XML parsing (no external dependencies)."""
    import io
    import zipfile

    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            with zf.open("word/document.xml") as xml_file:
                xml_content = xml_file.read()
    except zipfile.BadZipFile:
        return {"success": False, "error": "Invalid DOCX file (not a valid ZIP)"}
    except KeyError:
        return {"success": False, "error": "Invalid DOCX file (missing word/document.xml)"}

    try:
        import xml.etree.ElementTree as ET

        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        root = ET.fromstring(xml_content)

        paragraphs = []
        for para in root.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p"):
            texts = []
            for t in para.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t"):
                if t.text:
                    texts.append(t.text)
            para_text = "".join(texts).strip()
            if para_text:
                paragraphs.append(para_text)

        text = "\n".join(paragraphs)
        return {"success": True, "text": text, "page_count": 1, "file_type": "docx"}
    except ET.ParseError as e:
        return {"success": False, "error": f"Failed to parse DOCX XML: {e}"}
    except Exception as e:
        return {"success": False, "error": f"DOCX processing failed: {e}"}


def _process_xlsx(data: bytes) -> dict[str, Any]:
    """Process Excel spreadsheet."""
    import openpyxl

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(data)
        temp_path = f.name

    try:
        wb = openpyxl.load_workbook(temp_path, data_only=True)
        sheets_text = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = [str(cell.value or "") for row in ws.iter_rows() for cell in row]
            sheets_text.append(f"=== {sheet} ===\n" + "\n".join(rows))
        return {
            "success": True,
            "text": "\n\n".join(sheets_text),
            "page_count": len(wb.sheetnames),
            "file_type": "xlsx",
        }
    except ImportError:
        return {"success": False, "error": "openpyxl not installed"}
    except Exception as e:
        return {"success": False, "error": f"XLSX processing failed: {e}"}
    finally:
        os.unlink(temp_path)


def _process_zip(data: bytes) -> dict[str, Any]:
    """Process ZIP archive - extract file list only."""
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            names = zf.namelist()
        return {"success": True, "text": "\n".join(names), "page_count": 0, "file_type": "zip"}
    except Exception as e:
        return {"success": False, "error": f"ZIP processing failed: {e}"}


def _process_image(data: bytes) -> dict[str, Any]:
    """Process image - extract basic info."""
    return {"success": True, "text": f"[Image: {len(data)} bytes]", "page_count": 1, "file_type": "image"}
